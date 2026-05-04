from __future__ import annotations

import csv
import io
import json
from datetime import UTC, datetime
from typing import Any, cast

import httpx
import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]
from sqlalchemy.orm import Session
from starlette import status

from app.db.session import get_session_factory, initialize_metadata_database
from app.main import app
from app.models.connection import Connection
from app.models.label import LabelRecord, LabelSchema
from app.models.named_query import NamedQuery
from app.models.view_config import ViewConfig
from app.services.export_service import ExportService
from app.services.query_executor import Column, ExecutorResult, ExecutorService

HTTP_OK = status.HTTP_200_OK
DEFAULT_TIMEOUT = 30
DEFAULT_ROW_LIMIT = 10000


class FakeExecutorService(ExecutorService):
    def __init__(self, result: ExecutorResult) -> None:
        self.result = result
        self.calls: list[tuple[str, int, int]] = []

    def execute(
        self,
        connection: Connection,
        sql: str,
        *,
        timeout: int,
        row_limit: int,
    ) -> ExecutorResult:
        assert connection.id > 0
        self.calls.append((sql, timeout, row_limit))
        return self.result


def test_export_csv_includes_labels_and_bom() -> None:
    session, query_id, executor = _prepare_export_case()
    try:
        file_bytes, filename = ExportService(executor).export(
            session,
            query_id=query_id,
            format="csv",
        )
    finally:
        session.close()

    assert file_bytes.startswith("\ufeff".encode("utf-8"))
    assert filename.startswith("中文查询_")
    assert filename.endswith(".csv")
    assert executor.calls == [("SELECT * FROM trajectory_rows", 30, 10000)]

    rows = list(csv.reader(io.StringIO(file_bytes.decode("utf-8-sig"), newline="")))
    assert rows[0] == [
        "content",
        "id",
        "meta",
        "created_at",
        "score",
        "nullable",
        "label__quality",
        "label__tags",
        "label__note",
    ]
    assert rows[1] == [
        "你好",
        "r1",
        json.dumps({"tokens": 3, "tools": ["search"]}, ensure_ascii=False),
        "2026-01-02T03:04:05+00:00",
        "1.5",
        "",
        "良好",
        "安全|缺陷",
        "需要复查",
    ]
    assert rows[2][-3:] == ["不好", "", ""]
    assert "hidden_col" not in rows[0]


def test_export_xlsx_can_be_read_back() -> None:
    session, query_id, executor = _prepare_export_case()
    try:
        file_bytes, filename = ExportService(executor).export(
            session,
            query_id=query_id,
            format="xlsx",
        )
    finally:
        session.close()

    assert filename.startswith("中文查询_")
    assert filename.endswith(".xlsx")

    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == (
        "content",
        "id",
        "meta",
        "created_at",
        "score",
        "nullable",
        "label__quality",
        "label__tags",
        "label__note",
    )
    assert rows[1][0] == "你好"
    assert rows[1][-3:] == ("良好", "安全|缺陷", "需要复查")
    assert rows[2][-3:] == ("不好", None, None)
    workbook.close()


def test_export_creates_missing_view_config_for_legacy_query() -> None:
    session, query_id, executor = _prepare_export_case(with_view_config=False)
    try:
        file_bytes, _ = ExportService(executor).export(
            session,
            query_id=query_id,
            format="csv",
            include_labels=False,
        )
        view_config_count = session.query(ViewConfig).filter_by(query_id=query_id).count()
    finally:
        session.close()

    assert file_bytes.startswith("\ufeff".encode("utf-8"))
    assert view_config_count == 1


def test_export_escapes_spreadsheet_formula_values() -> None:
    session, query_id, executor = _prepare_formula_export_case()
    try:
        csv_bytes, _ = ExportService(executor).export(
            session,
            query_id=query_id,
            format="csv",
        )
        xlsx_bytes, _ = ExportService(executor).export(
            session,
            query_id=query_id,
            format="xlsx",
        )
    finally:
        session.close()

    csv_rows = list(csv.reader(io.StringIO(csv_bytes.decode("utf-8-sig"), newline="")))
    assert csv_rows[0] == ["'=alias", "safe", "label__'=label"]
    assert csv_rows[1] == ["'=SUM(1,1)", '\' \t@HYPERLINK("http://example.com")', "'+review"]
    assert csv_rows[2] == ["'-42", "plain", "'-risk"]

    workbook = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=False)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == ("'=alias", "safe", "label__'=label")
    assert rows[1] == ("'=SUM(1,1)", '\' \t@HYPERLINK("http://example.com")', "'+review")
    assert rows[2] == ("'-42", "plain", "'-risk")
    workbook.close()


def test_export_applies_saved_table_sort_and_keeps_labels_aligned() -> None:
    session, query_id, executor = _prepare_sorted_export_case()
    try:
        file_bytes, _ = ExportService(executor).export(
            session,
            query_id=query_id,
            format="csv",
        )
    finally:
        session.close()

    rows = list(csv.reader(io.StringIO(file_bytes.decode("utf-8-sig"), newline="")))
    assert rows[0] == ["id", "score", "content", "label__quality"]
    assert rows[1:] == [
        ["r2", "1", "low", "不好"],
        ["r3", "2", "middle", "良好"],
        ["r1", "3", "high", "良好"],
    ]


@pytest.mark.asyncio
async def test_export_api_returns_stream_with_encoded_filename(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    session, query_id, executor = _prepare_export_case()
    result = executor.result
    session.close()

    def fake_execute(
        self: ExecutorService,
        connection: Connection,
        sql: str,
        *,
        timeout: int,
        row_limit: int,
    ) -> ExecutorResult:
        assert isinstance(self, ExecutorService)
        assert connection.id > 0
        assert sql == "SELECT * FROM trajectory_rows"
        assert timeout == DEFAULT_TIMEOUT
        assert row_limit == DEFAULT_ROW_LIMIT
        return result

    monkeypatch.setattr(ExecutorService, "execute", fake_execute)

    transport = httpx.ASGITransport(app=cast(Any, app))
    async with httpx.AsyncClient(transport=transport, base_url="http://testserver") as client:
        response = await client.post(
            f"/api/v1/queries/{query_id}/export",
            json={"format": "csv", "include_labels": True, "json_serialization": "string"},
        )

    assert response.status_code == HTTP_OK
    assert response.headers["content-type"].startswith("text/csv")
    assert (
        "filename*=UTF-8''%E4%B8%AD%E6%96%87%E6%9F%A5%E8%AF%A2_"
        in response.headers["content-disposition"]
    )
    assert response.content.startswith("\ufeff".encode("utf-8"))


def _prepare_export_case(
    *, with_view_config: bool = True
) -> tuple[Session, int, FakeExecutorService]:
    initialize_metadata_database()
    session = get_session_factory()()
    connection = Connection(
        name="test-mysql",
        db_type="mysql",
        host="db.example.com",
        port=3306,
        database="agent_logs",
        username="reader",
        default_timeout=DEFAULT_TIMEOUT,
        default_row_limit=DEFAULT_ROW_LIMIT,
    )
    session.add(connection)
    session.flush()

    query = NamedQuery(
        connection_id=connection.id,
        name="中文查询",
        sql_text="SELECT * FROM trajectory_rows",
        is_named=True,
    )
    if with_view_config:
        query.view_config = ViewConfig(
            field_renders="{}",
            table_config=json.dumps(
                {
                    "hidden_columns": ["hidden_col"],
                    "column_order": ["content", "id", "meta", "created_at", "missing_col"],
                },
                ensure_ascii=False,
            ),
            row_identity_column="id",
        )
    query.label_schema = LabelSchema(
        fields=json.dumps(
            [
                {
                    "key": "quality",
                    "label": "质量",
                    "type": "single_select",
                    "options": [
                        {"value": "good", "label": "良好"},
                        {"value": "bad", "label": "不好"},
                    ],
                },
                {
                    "key": "tags",
                    "label": "标签",
                    "type": "multi_select",
                    "options": [
                        {"value": "safe", "label": "安全"},
                        {"value": "bug", "label": "缺陷"},
                    ],
                },
                {"key": "note", "label": "备注", "type": "text"},
            ],
            ensure_ascii=False,
        )
    )
    session.add(query)
    session.flush()
    session.add_all(
        [
            LabelRecord(
                query_id=query.id,
                row_identity="r1",
                field_key="quality",
                value=json.dumps("good"),
            ),
            LabelRecord(
                query_id=query.id,
                row_identity="r1",
                field_key="tags",
                value=json.dumps(["safe", "bug"]),
            ),
            LabelRecord(
                query_id=query.id,
                row_identity="r1",
                field_key="note",
                value=json.dumps("需要复查", ensure_ascii=False),
            ),
            LabelRecord(
                query_id=query.id,
                row_identity="r2",
                field_key="quality",
                value=json.dumps("bad"),
            ),
        ]
    )
    session.commit()

    result = ExecutorResult(
        columns=[
            Column(name="id", sql_type="VAR_STRING", inferred_type="text"),
            Column(name="content", sql_type="VAR_STRING", inferred_type="text"),
            Column(name="meta", sql_type="JSON", inferred_type="json"),
            Column(name="created_at", sql_type="DATETIME", inferred_type="timestamp"),
            Column(name="hidden_col", sql_type="VAR_STRING", inferred_type="text"),
            Column(name="score", sql_type="DOUBLE", inferred_type="float"),
            Column(name="nullable", sql_type="VAR_STRING", inferred_type="text"),
        ],
        rows=[
            {
                "id": "r1",
                "content": "你好",
                "meta": {"tokens": 3, "tools": ["search"]},
                "created_at": datetime(2026, 1, 2, 3, 4, 5, tzinfo=UTC),
                "hidden_col": "secret",
                "score": 1.5,
                "nullable": None,
            },
            {
                "id": "r2",
                "content": "bye",
                "meta": [],
                "created_at": "2026-01-03T00:00:00+00:00",
                "hidden_col": "secret",
                "score": 2,
                "nullable": None,
            },
        ],
        duration_ms=12,
        truncated=False,
    )
    return session, query.id, FakeExecutorService(result)


def _prepare_formula_export_case() -> tuple[Session, int, FakeExecutorService]:
    initialize_metadata_database()
    session = get_session_factory()()
    connection = Connection(
        name="test-mysql",
        db_type="mysql",
        host="db.example.com",
        port=3306,
        database="agent_logs",
        username="reader",
        default_timeout=DEFAULT_TIMEOUT,
        default_row_limit=DEFAULT_ROW_LIMIT,
    )
    session.add(connection)
    session.flush()

    query = NamedQuery(
        connection_id=connection.id,
        name="formula-risk",
        sql_text="SELECT formula columns",
        is_named=True,
    )
    query.view_config = ViewConfig(
        field_renders="{}",
        table_config=json.dumps({"hidden_columns": ["id"]}),
        row_identity_column="id",
    )
    query.label_schema = LabelSchema(
        fields=json.dumps(
            [
                {
                    "key": "'=label",
                    "label": "Danger",
                    "type": "single_select",
                    "options": [
                        {"value": "review", "label": "+review"},
                        {"value": "risk", "label": "-risk"},
                    ],
                }
            ]
        )
    )
    session.add(query)
    session.flush()
    session.add_all(
        [
            LabelRecord(
                query_id=query.id,
                row_identity="r1",
                field_key="'=label",
                value=json.dumps("review"),
            ),
            LabelRecord(
                query_id=query.id,
                row_identity="r2",
                field_key="'=label",
                value=json.dumps("risk"),
            ),
        ]
    )
    session.commit()

    result = ExecutorResult(
        columns=[
            Column(name="id", sql_type="VAR_STRING", inferred_type="text"),
            Column(name="=alias", sql_type="VAR_STRING", inferred_type="text"),
            Column(name="safe", sql_type="VAR_STRING", inferred_type="text"),
        ],
        rows=[
            {
                "id": "r1",
                "=alias": "=SUM(1,1)",
                "safe": ' \t@HYPERLINK("http://example.com")',
            },
            {
                "id": "r2",
                "=alias": -42,
                "safe": "plain",
            },
        ],
        duration_ms=12,
        truncated=False,
    )
    return session, query.id, FakeExecutorService(result)


def _prepare_sorted_export_case() -> tuple[Session, int, FakeExecutorService]:
    initialize_metadata_database()
    session = get_session_factory()()
    connection = Connection(
        name="test-mysql",
        db_type="mysql",
        host="db.example.com",
        port=3306,
        database="agent_logs",
        username="reader",
        default_timeout=DEFAULT_TIMEOUT,
        default_row_limit=DEFAULT_ROW_LIMIT,
    )
    session.add(connection)
    session.flush()

    query = NamedQuery(
        connection_id=connection.id,
        name="sorted-export",
        sql_text="SELECT sorted rows",
        is_named=True,
    )
    query.view_config = ViewConfig(
        field_renders="{}",
        table_config=json.dumps({"sort": [{"column": "score", "direction": "asc"}]}),
        row_identity_column="id",
    )
    query.label_schema = LabelSchema(
        fields=json.dumps(
            [
                {
                    "key": "quality",
                    "label": "Quality",
                    "type": "single_select",
                    "options": [
                        {"value": "good", "label": "良好"},
                        {"value": "bad", "label": "不好"},
                    ],
                }
            ],
            ensure_ascii=False,
        )
    )
    session.add(query)
    session.flush()
    session.add_all(
        [
            LabelRecord(
                query_id=query.id,
                row_identity="r1",
                field_key="quality",
                value=json.dumps("good"),
            ),
            LabelRecord(
                query_id=query.id,
                row_identity="r2",
                field_key="quality",
                value=json.dumps("bad"),
            ),
            LabelRecord(
                query_id=query.id,
                row_identity="r3",
                field_key="quality",
                value=json.dumps("good"),
            ),
        ]
    )
    session.commit()

    result = ExecutorResult(
        columns=[
            Column(name="id", sql_type="VAR_STRING", inferred_type="text"),
            Column(name="score", sql_type="LONGLONG", inferred_type="integer"),
            Column(name="content", sql_type="VAR_STRING", inferred_type="text"),
        ],
        rows=[
            {"id": "r1", "score": 3, "content": "high"},
            {"id": "r2", "score": 1, "content": "low"},
            {"id": "r3", "score": 2, "content": "middle"},
        ],
        duration_ms=12,
        truncated=False,
    )
    return session, query.id, FakeExecutorService(result)
